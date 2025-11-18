# ============================================================
# tests.test_xrepeat: comprehensive tests for xrepeat feature
# ============================================================

import pytest
from openpyxl import load_workbook
from pathlib import Path
from xml.etree import ElementTree as ET

from exlang import compile_xlang_to_xlsx, validate_xlang_minimal


# ============================================================
# Basic xrepeat Tests
# ============================================================

def test_xrepeat_basic_down(tmp_path):
    """Basic xrepeat with downward direction."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3" r="1" c="A">
          <xv>Row {{i}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_basic_down.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Row 1"
    assert ws["A2"].value == "Row 2"
    assert ws["A3"].value == "Row 3"


def test_xrepeat_basic_right(tmp_path):
    """Basic xrepeat with rightward direction."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3" r="1" c="A" direction="right">
          <xv>Col {{i}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_basic_right.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Col 1"
    assert ws["B1"].value == "Col 2"
    assert ws["C1"].value == "Col 3"


def test_xrepeat_multiple_xv(tmp_path):
    """xrepeat with multiple xv elements."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="4" r="2" c="B">
          <xv>Month {{i}}</xv>
          <xv>0</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_multiple_xv.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Row 2
    assert ws["B2"].value == "Month 1"
    assert ws["C2"].value == 0
    # Row 3
    assert ws["B3"].value == "Month 2"
    assert ws["C3"].value == 0
    # Row 4
    assert ws["B4"].value == "Month 3"
    assert ws["C4"].value == 0
    # Row 5
    assert ws["B5"].value == "Month 4"
    assert ws["C5"].value == 0


def test_xrepeat_zero_based_index(tmp_path):
    """xrepeat with {{i0}} zero-based index."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3" r="1" c="A">
          <xv>Index {{i0}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_i0.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Index 0"
    assert ws["A2"].value == "Index 1"
    assert ws["A3"].value == "Index 2"


def test_xrepeat_both_indices(tmp_path):
    """xrepeat with both {{i}} and {{i0}}."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="2" r="1" c="A">
          <xv>Row {{i}}</xv>
          <xv>Index {{i0}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_both_indices.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Row 1"
    assert ws["B1"].value == "Index 0"
    assert ws["A2"].value == "Row 2"
    assert ws["B2"].value == "Index 1"


# ============================================================
# Direction Tests
# ============================================================

def test_xrepeat_direction_down_explicit(tmp_path):
    """Explicit direction='down' works correctly."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3" r="5" c="C" direction="down">
          <xv>Item {{i}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_down_explicit.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["C5"].value == "Item 1"
    assert ws["C6"].value == "Item 2"
    assert ws["C7"].value == "Item 3"


def test_xrepeat_direction_right_multiple_xv(tmp_path):
    """direction='right' with multiple xv elements."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3" r="1" c="A" direction="right">
          <xv>Q{{i}}</xv>
          <xv>0</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_right_multiple.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Column A
    assert ws["A1"].value == "Q1"
    assert ws["A2"].value == 0
    # Column B
    assert ws["B1"].value == "Q2"
    assert ws["B2"].value == 0
    # Column C
    assert ws["C1"].value == "Q3"
    assert ws["C2"].value == 0


# ============================================================
# Default Values Tests
# ============================================================

def test_xrepeat_default_position(tmp_path):
    """xrepeat defaults to r=1, c=A."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="2">
          <xv>Default {{i}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_defaults.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Default 1"
    assert ws["A2"].value == "Default 2"


def test_xrepeat_default_direction(tmp_path):
    """xrepeat defaults to direction=down."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="2" r="3" c="B">
          <xv>Down {{i}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_default_direction.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["B3"].value == "Down 1"
    assert ws["B4"].value == "Down 2"


# ============================================================
# Integration Tests
# ============================================================

def test_xrepeat_with_xrow(tmp_path):
    """xrepeat combined with xrow."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrow r="1"><xv>Header</xv><xv>Value</xv></xrow>
        <xrepeat times="3" r="2" c="A">
          <xv>Row {{i}}</xv>
          <xv>{{i0}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_with_xrow.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Header
    assert ws["A1"].value == "Header"
    assert ws["B1"].value == "Value"
    # xrepeat data
    assert ws["A2"].value == "Row 1"
    assert ws["B2"].value == 0
    assert ws["A3"].value == "Row 2"
    assert ws["B3"].value == 1


def test_xrepeat_with_xcell_override(tmp_path):
    """xcell overrides xrepeat (last write wins)."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="5" r="1" c="A">
          <xv>Original {{i}}</xv>
        </xrepeat>
        <xcell addr="A3" v="Overridden"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_xcell_override.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Original 1"
    assert ws["A2"].value == "Original 2"
    assert ws["A3"].value == "Overridden"  # xcell wins
    assert ws["A4"].value == "Original 4"
    assert ws["A5"].value == "Original 5"


def test_xrepeat_with_xrange(tmp_path):
    """xrepeat works with xrange."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrange from="A1" to="A10" fill="Default"/>
        <xrepeat times="3" r="2" c="A">
          <xv>Repeat {{i}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_with_xrange.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Default"
    assert ws["A2"].value == "Repeat 1"  # xrepeat overwrites xrange
    assert ws["A3"].value == "Repeat 2"
    assert ws["A4"].value == "Repeat 3"
    assert ws["A5"].value == "Default"


def test_multiple_xrepeats(tmp_path):
    """Multiple xrepeat elements in same sheet."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="2" r="1" c="A">
          <xv>First {{i}}</xv>
        </xrepeat>
        <xrepeat times="2" r="1" c="C">
          <xv>Second {{i}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "multiple_xrepeats.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "First 1"
    assert ws["A2"].value == "First 2"
    assert ws["C1"].value == "Second 1"
    assert ws["C2"].value == "Second 2"


# ============================================================
# Large Repetition Tests
# ============================================================

def test_xrepeat_large_times(tmp_path):
    """xrepeat with large times value."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="50" r="1" c="A">
          <xv>Row {{i}}</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_large.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    assert ws["A1"].value == "Row 1"
    assert ws["A25"].value == "Row 25"
    assert ws["A50"].value == "Row 50"


def test_xrepeat_compression_benefit(tmp_path):
    """Demonstrate ORO benefit of xrepeat."""
    # This single xrepeat tag replaces 12 xrow tags
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrow r="1"><xv>Month</xv><xv>Budget</xv></xrow>
        <xrepeat times="12" r="2" c="A">
          <xv>Month {{i}}</xv>
          <xv>0</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "xrepeat_compression.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb["Test"]
    
    # Verify all 12 months created
    for i in range(1, 13):
        assert ws[f"A{i+1}"].value == f"Month {i}"
        assert ws[f"B{i+1}"].value == 0


# ============================================================
# Validation Error Tests
# ============================================================

def test_xrepeat_missing_times():
    """xrepeat without times attribute fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat r="1" c="A">
          <xv>Test</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 1
    assert "xrepeat missing required attribute 'times'" in errors[0]


def test_xrepeat_invalid_times_non_integer():
    """xrepeat with non-integer times fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="abc" r="1" c="A">
          <xv>Test</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 1
    assert "must be an integer" in errors[0]


def test_xrepeat_invalid_times_zero():
    """xrepeat with times=0 fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="0" r="1" c="A">
          <xv>Test</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 1
    assert "must be >= 1" in errors[0]


def test_xrepeat_invalid_times_negative():
    """xrepeat with negative times fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="-5" r="1" c="A">
          <xv>Test</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 1
    assert "must be >= 1" in errors[0]


def test_xrepeat_invalid_direction():
    """xrepeat with invalid direction fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3" direction="diagonal">
          <xv>Test</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 1
    assert "invalid direction" in errors[0]
    assert "diagonal" in errors[0]


def test_xrepeat_nested_not_allowed():
    """Nested xrepeat fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="2">
          <xv>Outer {{i}}</xv>
          <xrepeat times="2">
            <xv>Inner</xv>
          </xrepeat>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert any("Nested xrepeat is not allowed" in e for e in errors)


def test_xrepeat_invalid_content_xcell():
    """xrepeat with xcell child fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3">
          <xcell addr="A1" v="Test"/>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert any("can only contain <xv> tags" in e for e in errors)


def test_xrepeat_invalid_content_xrow():
    """xrepeat with xrow child fails validation."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xrepeat times="3">
          <xrow r="1"><xv>Test</xv></xrow>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert any("can only contain <xv> tags" in e for e in errors)
