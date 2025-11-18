# ============================================================
# tests.test_auto_naming: comprehensive tests for auto sheet naming
# ============================================================

import pytest
from openpyxl import load_workbook
from pathlib import Path
from xml.etree import ElementTree as ET

from exlang import compile_xlang_to_xlsx, validate_xlang_minimal


# ============================================================
# Basic Auto-Naming Tests
# ============================================================

def test_single_unnamed_sheet(tmp_path):
    """Single unnamed sheet should be named 'Sheet1'."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
    </xworkbook>
    """
    output = tmp_path / "single_unnamed.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert wb.sheetnames == ["Sheet1"]


def test_multiple_unnamed_sheets(tmp_path):
    """Multiple unnamed sheets should be numbered sequentially."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet></xsheet>
      <xsheet></xsheet>
    </xworkbook>
    """
    output = tmp_path / "multiple_unnamed.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert wb.sheetnames == ["Sheet1", "Sheet2", "Sheet3"]


def test_all_named_sheets(tmp_path):
    """Explicitly named sheets should use provided names."""
    xlang = """
    <xworkbook>
      <xsheet name="Data"></xsheet>
      <xsheet name="Summary"></xsheet>
    </xworkbook>
    """
    output = tmp_path / "all_named.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert wb.sheetnames == ["Data", "Summary"]


# ============================================================
# Mixed Naming Strategy Tests
# ============================================================

def test_mixed_named_and_unnamed(tmp_path):
    """Mixed named and unnamed sheets should work correctly."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet name="Data"></xsheet>
      <xsheet></xsheet>
    </xworkbook>
    """
    output = tmp_path / "mixed_naming.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert wb.sheetnames == ["Sheet1", "Data", "Sheet2"]


def test_unnamed_followed_by_named(tmp_path):
    """Unnamed sheets followed by named sheets."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet></xsheet>
      <xsheet name="Summary"></xsheet>
      <xsheet name="Report"></xsheet>
    </xworkbook>
    """
    output = tmp_path / "unnamed_then_named.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert wb.sheetnames == ["Sheet1", "Sheet2", "Summary", "Report"]


def test_named_followed_by_unnamed(tmp_path):
    """Named sheets followed by unnamed sheets."""
    xlang = """
    <xworkbook>
      <xsheet name="Data"></xsheet>
      <xsheet name="Summary"></xsheet>
      <xsheet></xsheet>
      <xsheet></xsheet>
    </xworkbook>
    """
    output = tmp_path / "named_then_unnamed.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert wb.sheetnames == ["Data", "Summary", "Sheet1", "Sheet2"]


# ============================================================
# Collision Detection Tests
# ============================================================

def test_collision_sheet1_explicit(tmp_path):
    """Collision when 'Sheet1' is explicitly named and unnamed sheet exists."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet name="Sheet1"></xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 1
    assert "Sheet1" in errors[0]
    assert "conflicts with explicitly named sheet" in errors[0]


def test_collision_sheet2_explicit(tmp_path):
    """Collision when 'Sheet2' is explicitly named and multiple unnamed sheets exist."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet></xsheet>
      <xsheet name="Sheet2"></xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 1
    assert "Sheet2" in errors[0]
    assert "conflicts with explicitly named sheet" in errors[0]


def test_collision_multiple_conflicts(tmp_path):
    """Multiple collision errors when multiple auto-generated names conflict."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet></xsheet>
      <xsheet></xsheet>
      <xsheet name="Sheet1"></xsheet>
      <xsheet name="Sheet3"></xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 2
    assert any("Sheet1" in e for e in errors)
    assert any("Sheet3" in e for e in errors)


def test_no_collision_different_names(tmp_path):
    """No collision when explicit names don't use Sheet1, Sheet2 pattern."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet></xsheet>
      <xsheet name="Data"></xsheet>
      <xsheet name="Summary"></xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 0


def test_no_collision_all_unnamed(tmp_path):
    """No collision when all sheets are unnamed."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet></xsheet>
      <xsheet></xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 0


def test_no_collision_all_named(tmp_path):
    """No collision when all sheets are explicitly named."""
    xlang = """
    <xworkbook>
      <xsheet name="Data"></xsheet>
      <xsheet name="Summary"></xsheet>
      <xsheet name="Report"></xsheet>
    </xworkbook>
    """
    root = ET.fromstring(xlang)
    errors = validate_xlang_minimal(root)
    
    assert len(errors) == 0


# ============================================================
# Integration Tests with Content
# ============================================================

def test_auto_named_sheet_with_content(tmp_path):
    """Auto-named sheet should correctly contain data."""
    xlang = """
    <xworkbook>
      <xsheet>
        <xcell addr="A1" v="Test Value"/>
        <xcell addr="B2" v="42" t="number"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "auto_named_with_content.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert wb.sheetnames == ["Sheet1"]
    
    ws = wb["Sheet1"]
    assert ws["A1"].value == "Test Value"
    assert ws["B2"].value == 42


def test_mixed_naming_with_xrow_and_xrange(tmp_path):
    """Mixed naming with xrow and xrange content."""
    xlang = """
    <xworkbook>
      <xsheet>
        <xrow r="1">
          <xv>Header1</xv>
          <xv>Header2</xv>
        </xrow>
      </xsheet>
      <xsheet name="Data">
        <xrange from="A1" to="A3" fill="0" t="number"/>
      </xsheet>
      <xsheet>
        <xcell addr="A1" v="Summary"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "mixed_with_content.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert wb.sheetnames == ["Sheet1", "Data", "Sheet2"]
    
    ws1 = wb["Sheet1"]
    assert ws1["A1"].value == "Header1"
    assert ws1["B1"].value == "Header2"
    
    ws2 = wb["Data"]
    assert ws2["A1"].value == 0
    assert ws2["A2"].value == 0
    assert ws2["A3"].value == 0
    
    ws3 = wb["Sheet2"]
    assert ws3["A1"].value == "Summary"


# ============================================================
# Edge Case Tests
# ============================================================

def test_single_unnamed_sheet_minimal(tmp_path):
    """Minimal valid document with single unnamed sheet."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
    </xworkbook>
    """
    output = tmp_path / "minimal_unnamed.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "Sheet1"


def test_many_unnamed_sheets(tmp_path):
    """Large number of unnamed sheets."""
    sheets = "\n".join(["      <xsheet></xsheet>"] * 10)
    xlang = f"""
    <xworkbook>
{sheets}
    </xworkbook>
    """
    output = tmp_path / "many_unnamed.xlsx"
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    expected = [f"Sheet{i}" for i in range(1, 11)]
    assert wb.sheetnames == expected


def test_collision_compile_raises_error(tmp_path):
    """Compile should raise error when collision detected."""
    xlang = """
    <xworkbook>
      <xsheet></xsheet>
      <xsheet name="Sheet1"></xsheet>
    </xworkbook>
    """
    output = tmp_path / "collision_error.xlsx"
    
    with pytest.raises(ValueError, match="Sheet1.*conflicts"):
        compile_xlang_to_xlsx(xlang, output)
