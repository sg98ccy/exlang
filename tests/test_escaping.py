"""
Test automatic XML escaping for EXLang formulas.

The compiler automatically escapes XML special characters in formulas.
This allows natural formula syntax without manual escaping.
"""

from exlang import compile_xlang_to_xlsx
from openpyxl import load_workbook
from pathlib import Path


def test_jinja_basic_formula(tmp_path):
    """Test basic formula with < and quotes using inline syntax."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v='=IF(B1<100,"Low","High")'/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    # Formula is automatically escaped during compilation
    compile_xlang_to_xlsx(xlang, output)
    
    # Verify the file was created and formula is correct
    wb = load_workbook(output)
    ws = wb.active
    assert ws['A1'].value == '=IF(B1<100,"Low","High")'


def test_jinja_multiple_formulas(tmp_path):
    """Test multiple formulas with different operators."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v='=IF(B1<100,"Low","High")'/>
        <xcell addr="A2" v='=IF(B2>=100,"High","Low")'/>
        <xcell addr="A3" v='=IF(B3<>100,"Not Equal","Equal")'/>
        <xcell addr="A4" v="=A1&A2"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    # All formulas are automatically escaped during compilation
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb.active
    assert ws['A1'].value == '=IF(B1<100,"Low","High")'
    assert ws['A2'].value == '=IF(B2>=100,"High","Low")'
    assert ws['A3'].value == '=IF(B3<>100,"Not Equal","Equal")'
    assert ws['A4'].value == '=A1&A2'


def test_jinja_loop_formulas(tmp_path):
    """Test inline formulas in xrepeat with comparison operators."""
    xlang = """
    <xworkbook>
      <xsheet name="Inventory">
        <xrepeat times="3" r="4" c="K">
          <xv>=IF(J4<100,"REORDER","OK")</xv>
        </xrepeat>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb.active
    # All three rows get the same formula (no {{i}} substitution)
    assert ws['K4'].value == '=IF(J4<100,"REORDER","OK")'
    assert ws['K5'].value == '=IF(J4<100,"REORDER","OK")'
    assert ws['K6'].value == '=IF(J4<100,"REORDER","OK")'


def test_jinja_complex_nested_formula(tmp_path):
    """Test complex nested formulas with multiple operators."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v='=IF(AND(B1>50,B1<100),"Medium",IF(B1>=100,"High","Low"))'/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb.active
    assert ws['A1'].value == '=IF(AND(B1>50,B1<100),"Medium",IF(B1>=100,"High","Low"))'


def test_plain_xml_with_manual_escaping(tmp_path):
    """Test that plain XML with manual escaping still works (backward compatibility)."""
    xlang = """
    <xworkbook>
      <xsheet name="Test">
        <xcell addr="A1" v="=IF(B1&lt;100,&quot;Low&quot;,&quot;High&quot;)"/>
      </xsheet>
    </xworkbook>
    """
    output = tmp_path / "test.xlsx"
    
    # Manual escaping still works with automatic Jinja2 preprocessing
    compile_xlang_to_xlsx(xlang, output)
    
    wb = load_workbook(output)
    ws = wb.active
    assert ws['A1'].value == '=IF(B1<100,"Low","High")'
